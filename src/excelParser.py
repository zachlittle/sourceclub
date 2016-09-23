from openpyxl import load_workbook
import glob, os, shutil
import datetime
import csv
from openpyxl import Workbook

global mostRecentExcelFile
global currentSheet
global excelFileDir
productData = {}

CONSTRAINTS = {'Price': {'Min': 20, 'Max': 100},
               'Fees': {},
               'Net': {'Min': 10},
               'Weight': {'Max': 10},
               'Reviews': {},
               'AverageRating': {},
               'Rank': {},
               'EstimatedMonthlySales': {'Min': 5},
               'EstimatedMonthlyRevenue': {},
               'NumberOfSellers': {'Max': 15}
              }

class excelHandler:

    def __init__(self):
        global excelFileDir
        excelFileDir = 'C:\\Git\\sourceclub\\src\\Excel Files\\'
        self.findAndMoveExcelFiles()
        self.renameCSVFilesToXLSX()
        self.findMostRecentFileBasedOnMT()
        self.selectSheetByFilePath()
        #self.packageAllFilesRelatedIntoOneWorkbook()

    @staticmethod
    def selectSheetByFilePath():
        global currentSheet
        excelFileName = excelFileDir + str(mostRecentExcelFile)
        wb2 = load_workbook(excelFileName)
        sheetName = wb2.get_sheet_names()
        currentSheet = wb2.get_sheet_by_name(sheetName[0])
        # sheetToListGenerator(currentSheet)

    @staticmethod
    def findMostRecentFileBasedOnMT():
        import os
        guyList = []
        maximumModTime = 0
        [guyList.append(os.path.getmtime(excelFileDir + "" + str(file))) for file in os.listdir(excelFileDir) if file.endswith(".xlsx")]
        fileModTimes = [modTime for modTime in guyList if modTime > maximumModTime]
        [float(modTime) for modTime in fileModTimes]
        maximumModTime = max(fileModTimes)
        dt_obj = datetime.datetime.utcfromtimestamp(maximumModTime)
        print("Most Recent Excel File To Parse: " + str(dt_obj))
        excelHandler.selectMostRecentFileBasedOnMT(dt_obj)

    @staticmethod
    def selectMostRecentFileBasedOnMT(dateTime):
        global mostRecentExcelFile
        for file in os.listdir(excelFileDir):
            print("----------------------------------------------------------------------------------")
            currentFileDateTime = datetime.datetime.utcfromtimestamp(os.path.getmtime(excelFileDir + "" + str(file)))
            if currentFileDateTime == dateTime:
                print("Most Recent File: " + file)
                mostRecentExcelFile = file

    def findAndMoveExcelFiles(self):
        source_dir = 'C:\\Users\maxsm_000\Downloads'
        [self.copyFilesToExcelDir(file, source_dir) for file in os.listdir(source_dir) if file.endswith(".csv")]

    @staticmethod
    def copyFilesToExcelDir(file, source_dir):
        shutil.copy2(source_dir + '\\' + file, excelFileDir)
        os.remove(source_dir + '\\' + file)

    def renameCSVFilesToXLSX(self):
        [TransferExcelData.ConvertCSVToXLSX(file) for file in os.listdir(excelFileDir) if file.endswith(".csv")]

class TransferExcelData:
    currentFileParsing = ""
    #TODO: Will handle all renaming of csv files as it requires full data reentry not just ext. rename.
    #TODO: Will handle merging and extraction of multiple flat files

    @staticmethod
    def ConvertCSVToXLSX(file):
        workBook = TransferExcelData.stripAndReEnterWBData(file)
        workBook.save(excelFileDir + str(TransferExcelData.uniqueStampGenerator()) + ".xlsx")
        TransferExcelData.deleteOldCSVFile()

    @staticmethod
    def deleteOldCSVFile():
        os.remove(excelFileDir + str(TransferExcelData.currentFileParsing))


    @staticmethod
    def uniqueStampGenerator():
        currentTime = datetime.datetime.now().time()
        newFileName = "productResearch_" + str(TransferExcelData.removeNonAlphaNumeric(currentTime))
        return newFileName

    @staticmethod
    def removeNonAlphaNumeric(string):
        modifiedString = ""
        for i, char in enumerate(str(string)):
            if char.isalpha() or char.isdigit():
                modifiedString += char
        return modifiedString

    @staticmethod
    def stripAndReEnterWBData(file):
        TransferExcelData.currentFileParsing = file
        wb = Workbook()
        sheet = wb.active
        excelFile = excelFileDir + str(file)
        CSV_SEPARATOR = ","
        with open(excelFile) as f:
            reader = csv.reader(f)
            for r, row in enumerate(reader):
                for c, col in enumerate(row):
                    for idx, val in enumerate(col.split(CSV_SEPARATOR)):
                        cell = sheet.cell(row=r + 1, column=c + 1)
                        cell.value = val
        return wb


class sheetToListGenerator:
    primaryKeyColumn = 'A'
    primaryKeyList = {}
    filteredProducts = []

    def __init__(self, sheet):
        self.createPrimaryKeyBasedOnColumn(sheet)
        self.populateDataStructure()
        self.outputStructure()
        self.filterByConstraints()
        self.printStats()

    def printStats(self):
        print('--------------------FILTERSTATS---------------------')
        print('# of products filtered due to failure to meet constraints: ' + str(len(self.filteredProducts)))
        print("Name's of products filtered: ")
        for product in self.filteredProducts:
            print(product)
        print('# of viable products remaining: ' + str(len(productData)))
        print("Name's of products remaining: ")
        for product in productData:
            print(product)
        print('--------------------FILTERSTATS---------------------')

    def filterByConstraints(self):
        productsToRemove = []
        for keyIndex, key in enumerate(self.primaryKeyList):
            for category in productData[key]:
                if CONSTRAINTS.get(str(category)) is not None:
                    if type(productData[key][str(category)]) is not str:
                        if 'Min' in CONSTRAINTS[str(category)].keys():
                            if productData[key][str(category)] < CONSTRAINTS[str(category)]['Min']:
                                productsToRemove.append(key)
                                break
                        if 'Max' in CONSTRAINTS[str(category)].keys():
                            if productData[key][str(category)] > CONSTRAINTS[str(category)]['Max']:
                                productsToRemove.append(key)
                                break
        self.removeProductsByIndex(productsToRemove)

    def removeProductsByIndex(self, productsToRemove):
        for product in productsToRemove:
            self.filteredProducts.append(product)
            del productData[product]
            del self.primaryKeyList[product]

    def createPrimaryKeyBasedOnColumn(self, sheet):
        for i, row in enumerate(range(4, sheet.max_row + 1)):
            primaryKey = currentSheet['' + self.primaryKeyColumn + '' + str(row)].value
            productData.setdefault('' + primaryKey + '', {})
            self.primaryKeyList.setdefault('' + primaryKey + '', {})

    def populateDataStructure(self):
        self.createDataItem('Name', 'B')
        self.createDataItem('Price', 'F')
        self.createDataItem('Brand', 'C')
        self.createDataItem('Seller', 'D')
        self.createDataItem('Category', 'E')
        self.createDataItem('Fees', 'G')
        self.createDataItem('Net', 'H')
        self.createDataItem('Weight', 'I')
        self.createDataItem('ProductTier', 'J')
        self.createDataItem('Reviews', 'K')
        self.createDataItem('AverageRating', 'L')
        self.createDataItem('Rank', 'M')
        self.createDataItem('EstimatedMonthlySales', 'N')
        self.createDataItem('EstimatedMonthlyRevenue', 'O')
        self.createDataItem('NumberOfSellers', 'Q')

    @staticmethod
    def createDataItem(itemToAdd, columnValue):
        for row in range(4, currentSheet.max_row + 1):
            primaryKey = currentSheet['' + sheetToListGenerator.primaryKeyColumn + '' + str(row)].value
            itemValues = currentSheet['' + columnValue + '' + str(row)].value
            productData[primaryKey].setdefault(str(itemToAdd), itemValues)

    @staticmethod
    def outputStructure():
        for i, key in enumerate(sheetToListGenerator.primaryKeyList):
            print('--------------Item' + str(i) + '------------------')
            print('ASIN: ' + key)
            print('Name: ' + productData[key]['Name'])
            print('Brand: ' + str(productData[key]['Brand']))
            print('Price: ' + str(productData[key]['Price']))
            print('Seller: ' + str(productData[key]['Seller']))
            print('Category: ' + str(productData[key]['Category']))
            print('Fees: ' + str(productData[key]['Fees']))
            print('Net: ' + str(productData[key]['Net']))
            print('Weight: ' + str(productData[key]['Weight']))
            print('ProductTier: ' + str(productData[key]['ProductTier']))
            print('Reviews: ' + str(productData[key]['Reviews']))
            print('AverageRating: ' + str(productData[key]['AverageRating']))
            print('Rank: ' + str(productData[key]['Rank']))
            print('EstimatedMonthlySales: ' + str(productData[key]['EstimatedMonthlySales']))
            print('EstimatedMonthlyRevenue: ' + str(productData[key]['EstimatedMonthlyRevenue']))
            print('NumberOfSellers: ' + str(productData[key]['NumberOfSellers']))

if __name__ == "__main__":
    excelHandler()
# TODO: I may not want this below
excelHandler()
